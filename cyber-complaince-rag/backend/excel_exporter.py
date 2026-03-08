"""
Export query results to a formatted Excel workbook.
Matches the Unified Risk and Control Framework output style.
"""

from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─── Color palette (matching screenshots) ────────────────────────────────────
BLACK_HEADER  = "FF000000"
PURPLE_HEADER = "FF6B0080"   # dark purple
NAVY_HEADER   = "FF1B2B6B"   # dark navy
LIGHT_GRAY    = "FFF2F2F2"
WHITE         = "FFFFFFFF"
LIGHT_PURPLE  = "FFE8D5F5"
TEAL_ACCENT   = "FF14B8A6"


def _header_cell(ws, row, col, value,
                 bg=NAVY_HEADER, fg=WHITE,
                 bold=True, wrap=True, font_size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=font_size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
    return cell


def _data_cell(ws, row, col, value, bg=WHITE, bold=False,
               wrap=True, align="left", font_size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color="FF000000", size=font_size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="top",
                                wrap_text=wrap)
    return cell


def _thin_border():
    s = Side(style="thin", color="FFCCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def export_results_to_excel(
    results: list[dict],
    query: str,
    analysis: str,
    output_path: str,
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title row
    ws_summary.merge_cells("A1:H1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Unified Risk and Control Framework — Compliance Query Results"
    title_cell.font = Font(bold=True, size=14, color=WHITE, name="Arial")
    title_cell.fill = PatternFill("solid", fgColor=BLACK_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 32

    # Meta rows
    ws_summary.merge_cells("A2:H2")
    ws_summary["A2"].value = f"Query: {query}"
    ws_summary["A2"].font = Font(bold=True, size=10, name="Arial")
    ws_summary["A2"].fill = PatternFill("solid", fgColor=LIGHT_PURPLE)
    ws_summary["A2"].alignment = Alignment(horizontal="left", vertical="center",
                                            indent=1)
    ws_summary.row_dimensions[2].height = 20

    ws_summary.merge_cells("A3:H3")
    ws_summary["A3"].value = f"AI Analysis: {analysis}"
    ws_summary["A3"].font = Font(size=9, italic=True, name="Arial")
    ws_summary["A3"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws_summary["A3"].alignment = Alignment(horizontal="left", vertical="top",
                                            wrap_text=True, indent=1)
    ws_summary.row_dimensions[3].height = 60

    ws_summary.merge_cells("A4:H4")
    ws_summary["A4"].value = f"Total Results: {len(results)}"
    ws_summary["A4"].font = Font(size=9, name="Arial")
    ws_summary["A4"].alignment = Alignment(horizontal="left", indent=1)
    ws_summary.row_dimensions[4].height = 16

    # Section header spanning both panels
    ws_summary.merge_cells("A5:H5")
    sec_cell = ws_summary["A5"]
    sec_cell.value = "Matched Requirements"
    sec_cell.font = Font(bold=True, size=11, color=WHITE, name="Arial")
    sec_cell.fill = PatternFill("solid", fgColor=PURPLE_HEADER)
    sec_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[5].height = 24

    # Column headers
    col_headers = ["#", "Framework", "Topic", "Sub Topic",
                   "Section Number", "Requirements", "Theme", "Relevance"]
    col_widths   = [4,   18,           20,       25,
                    16,               55,              30,     11]

    for col_i, (hdr, width) in enumerate(zip(col_headers, col_widths), start=1):
        _header_cell(ws_summary, 6, col_i, hdr, bg=NAVY_HEADER, font_size=9)
        ws_summary.column_dimensions[get_column_letter(col_i)].width = width

    ws_summary.row_dimensions[6].height = 20

    # Data rows
    for row_i, r in enumerate(results, start=7):
        alt_bg = LIGHT_GRAY if (row_i % 2 == 0) else WHITE
        _data_cell(ws_summary, row_i, 1, r.get("rank", row_i - 6),
                   bg=alt_bg, align="center")
        _data_cell(ws_summary, row_i, 2, r.get("framework", ""), bg=alt_bg)
        _data_cell(ws_summary, row_i, 3, r.get("topic", ""), bg=alt_bg)
        _data_cell(ws_summary, row_i, 4, r.get("sub_topic", ""), bg=alt_bg)
        _data_cell(ws_summary, row_i, 5, r.get("section_number", ""),
                   bg=alt_bg, align="center")
        _data_cell(ws_summary, row_i, 6, r.get("requirements", ""), bg=alt_bg)
        _data_cell(ws_summary, row_i, 7, r.get("theme", ""), bg=alt_bg)
        score = r.get("relevance_score", "")
        score_cell = _data_cell(ws_summary, row_i, 8, score,
                                 bg=alt_bg, align="center", bold=True)
        # Color-code relevance score
        if isinstance(score, int):
            if score >= 8:
                score_cell.font = Font(bold=True, color="FF15803D", name="Arial")
            elif score >= 5:
                score_cell.font = Font(bold=True, color="FFB45309", name="Arial")
            else:
                score_cell.font = Font(bold=True, color="FFB91C1C", name="Arial")

        # Apply border
        for col_j in range(1, 9):
            ws_summary.cell(row=row_i, column=col_j).border = _thin_border()

        # Auto row height for long requirements
        req_len = len(r.get("requirements", ""))
        ws_summary.row_dimensions[row_i].height = max(18, min(req_len // 4, 80))

    ws_summary.freeze_panes = "A7"

    # ── Sheet 2: Per-Framework sheets ────────────────────────────────────────
    from collections import defaultdict
    by_fw: dict[str, list] = defaultdict(list)
    for r in results:
        by_fw[r.get("framework", "Unknown")].append(r)

    for fw_name, fw_results in sorted(by_fw.items()):
        safe_name = fw_name[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe_name)

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"].value = fw_name
        ws["A1"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
        ws["A1"].fill = PatternFill("solid", fgColor=NAVY_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"].value = f"Query Context: {query}  |  {len(fw_results)} requirements matched"
        ws["A2"].font = Font(size=9, italic=True, name="Arial")
        ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_PURPLE)
        ws["A2"].alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[2].height = 16

        headers = ["#", "Topic", "Sub Topic", "Section Number",
                   "Requirements", "Theme", "Relevance"]
        widths   = [4,   22,       26,           18,
                    58,            32,      11]

        for col_i, (hdr, width) in enumerate(zip(headers, widths), start=1):
            _header_cell(ws, 3, col_i, hdr, bg=PURPLE_HEADER, font_size=9)
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[3].height = 20

        for row_i, r in enumerate(fw_results, start=4):
            alt_bg = LIGHT_GRAY if (row_i % 2 == 0) else WHITE
            _data_cell(ws, row_i, 1, r.get("rank", ""), bg=alt_bg, align="center")
            _data_cell(ws, row_i, 2, r.get("topic", ""), bg=alt_bg)
            _data_cell(ws, row_i, 3, r.get("sub_topic", ""), bg=alt_bg)
            _data_cell(ws, row_i, 4, r.get("section_number", ""),
                       bg=alt_bg, align="center")
            _data_cell(ws, row_i, 5, r.get("requirements", ""), bg=alt_bg)
            _data_cell(ws, row_i, 6, r.get("theme", ""), bg=alt_bg)
            _data_cell(ws, row_i, 7, r.get("relevance_score", ""),
                       bg=alt_bg, align="center", bold=True)

            for col_j in range(1, 8):
                ws.cell(row=row_i, column=col_j).border = _thin_border()

            req_len = len(r.get("requirements", ""))
            ws.row_dimensions[row_i].height = max(18, min(req_len // 4, 80))

        ws.freeze_panes = "A4"

    # ── Sheet 3: Theme Mapping ───────────────────────────────────────────────
    ws_themes = wb.create_sheet(title="Theme Mapping")
    ws_themes.merge_cells("A1:E1")
    ws_themes["A1"].value = "Theme Distribution Across Results"
    ws_themes["A1"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws_themes["A1"].fill = PatternFill("solid", fgColor=BLACK_HEADER)
    ws_themes["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_themes.row_dimensions[1].height = 28

    theme_headers = ["Theme", "Framework", "Count", "Top Section Numbers", "Sample Requirement"]
    theme_widths  = [35,       20,           8,        30,                    55]
    for col_i, (hdr, width) in enumerate(zip(theme_headers, theme_widths), start=1):
        _header_cell(ws_themes, 2, col_i, hdr, bg=NAVY_HEADER, font_size=9)
        ws_themes.column_dimensions[get_column_letter(col_i)].width = width
    ws_themes.row_dimensions[2].height = 20

    # Group results by theme
    from collections import defaultdict
    theme_groups: dict[str, list] = defaultdict(list)
    for r in results:
        theme_groups[r.get("theme", "Unknown")].append(r)

    for row_i, (theme, t_results) in enumerate(
        sorted(theme_groups.items(), key=lambda x: -len(x[1])), start=3
    ):
        alt_bg = LIGHT_GRAY if (row_i % 2 == 0) else WHITE
        frameworks = ", ".join(sorted(set(r.get("framework", "") for r in t_results)))
        sections = ", ".join(sorted(set(
            r.get("section_number", "") for r in t_results if r.get("section_number")
        ))[:5])
        sample_req = t_results[0].get("requirements", "")[:200] if t_results else ""

        _data_cell(ws_themes, row_i, 1, theme, bg=alt_bg, bold=True)
        _data_cell(ws_themes, row_i, 2, frameworks, bg=alt_bg)
        _data_cell(ws_themes, row_i, 3, len(t_results), bg=alt_bg, align="center")
        _data_cell(ws_themes, row_i, 4, sections, bg=alt_bg)
        _data_cell(ws_themes, row_i, 5, sample_req, bg=alt_bg)

        for col_j in range(1, 6):
            ws_themes.cell(row=row_i, column=col_j).border = _thin_border()
        ws_themes.row_dimensions[row_i].height = 20

    ws_themes.freeze_panes = "A3"

    # Save
    wb.save(output_path)
    return output_path
